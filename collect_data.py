#!/usr/bin/env python3
import os, json, sys, time, re
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError
from urllib.parse import urlencode
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "firework_video_data.xlsx"
MAX_RESULTS_PER_QUERY = 25
API_KEY = os.environ.get("YOUTUBE_API_KEY", "")

# Tier 1: high precision, near-zero noise
# Tier 2: medium precision, low noise
# Tier 3: extended reach, moderate noise (rely on filters)
SEARCH_QUERIES = [
    # Tier 1 — spec keywords that professional shows never use
    '"500g cake" firework',
    '"200g cake" firework',
    '"60g" firework fountain OR cake',
    '"ground fountain" OR "cone fountain" firework',
    '"repeater cake" OR "barrage cake" firework',
    # Tier 2 — consumer-level qualifiers
    "consumer firework cake demo OR test OR review OR unboxing",
    'backyard firework cake OR fountain demo OR test',
    '"multi shot" cake firework demo OR test OR review',
    '"compound cake" OR "aerial cake" firework demo OR review',
    # Tier 3 — extended reach with exclusion hints in filter
    "firework cake unboxing OR \"first look\" OR \"new for\"",
    "firework fountain review OR test -\"firework show\" -display -festival -\"4th of july\" -professional",
    "cake firework \"test fire\" OR \"shoot\" OR \"light it up\" -show -display -festival -professional -\"new year\"",
]

INCLUDE_KEYWORDS = [
    "cake", "barrage", "repeater", "multi-shot", "multi shot",
    "fountain", "ground fountain", "aerial cake",
    "firework cake", "firework fountain", "500g cake", "200g cake",
    "consumer firework", "backyard firework", "cake demo",
    "cone fountain", "ice fountain", "compound cake",
]

EXCLUDE_KEYWORDS = [
    "firework show", "fireworks show", "display shell",
    "professional display", "pyromusical", "firework competition",
    "firework festival", "music video", "how to make",
    "tutorial", "diy firework", "manufacturing", "drone",
    "wedding", "new year countdown", "new years eve",
    "4th of july show", "fourth of july show",
    "independence day show", "mall show", "stadium",
    "orchestra", "concert",
]

# Auto-fill powder weight (grams) when spec keyword found in title/description
POWDER_AUTO_FILL = {
    "500g": 500,
    "500 g": 500,
    "200g": 200,
    "200 g": 200,
    "350g": 350,
    "350 g": 350,
    "120g": 120,
    "120 g": 120,
    "60g": 60,
    "60 g": 60,
    "30g": 30,
    "30 g": 30,
}

# Strong consumer-level signals — boost relevance priority
STRONG_CONSUMER_SIGNALS = [
    "backyard", "consumer", "unboxing", "haul", "pickup",
    "500g", "200g", "60g", "1.4g",
    "demo", "test fire", "first look", "new for 202",
    "review", "shoot", "light it up", "cake", "fountain",
]

HEADERS = [
    "样本编号", "视频URL", "UP主名称", "UP主类型",
    "发布年份", "发布月份", "产品大类",
    "发数", "筒径规格", "药量(g)", "燃放时长", "效果类型",
    "播放量", "点击率CTR", "完播率",
    "点赞数", "评论数", "收藏数", "综合互动率",
]

MANUAL_COLS = {4, 8, 9, 10, 11, 12}
UNAVAILABLE_COLS = {14, 15, 18}


def _api_call(endpoint, params):
    params["key"] = API_KEY
    url = f"https://www.googleapis.com/youtube/v3/{endpoint}?{urlencode(params)}"
    proxy = os.environ.get("HTTPS_PROXY") or os.environ.get("https_proxy") or None
    try:
        if proxy:
            import ssl
            ctx = ssl.create_default_context()
            req = Request(url, headers={"User-Agent": "FireworkCollector/1.0"})
            req.set_proxy(proxy, "https")
            resp = urlopen(req, timeout=30, context=ctx)
        else:
            req = Request(url, headers={"User-Agent": "FireworkCollector/1.0"})
            resp = urlopen(req, timeout=30)
        return json.loads(resp.read().decode())
    except URLError as e:
        print(f"    API error: {e}")
        return None


def search_youtube_api(query, max_results=25):
    results = []
    page_token = None
    while len(results) < max_results:
        params = {
            "part": "snippet", "q": query, "type": "video",
            "maxResults": min(50, max_results - len(results)),
            "relevanceLanguage": "en", "regionCode": "US",
        }
        if page_token:
            params["pageToken"] = page_token
        data = _api_call("search", params)
        if not data or "items" not in data:
            break
        for item in data["items"]:
            vid = item["id"].get("videoId")
            if not vid:
                continue
            s = item.get("snippet", {})
            results.append({
                "id": vid,
                "title": s.get("title", ""),
                "description": s.get("description", ""),
                "uploader": s.get("channelTitle", ""),
                "published_at": s.get("publishedAt", ""),
                "webpage_url": f"https://www.youtube.com/watch?v={vid}",
            })
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    return results[:max_results]


def get_batch_details(video_ids):
    params = {
        "part": "statistics,contentDetails,snippet",
        "id": ",".join(video_ids), "maxResults": "50",
    }
    data = _api_call("videos", params)
    if not data or "items" not in data:
        return {}
    details = {}
    for item in data["items"]:
        vid = item["id"]
        s = item.get("snippet", {})
        st = item.get("statistics", {})
        c = item.get("contentDetails", {})
        details[vid] = {
            "title": s.get("title", ""),
            "description": s.get("description", ""),
            "uploader": s.get("channelTitle", ""),
            "published_at": s.get("publishedAt", ""),
            "view_count": int(st.get("viewCount", 0)),
            "like_count": int(st.get("likeCount", 0)),
            "comment_count": int(st.get("commentCount", 0)),
            "duration_iso": c.get("duration", ""),
        }
    return details


def parse_iso_duration(iso_dur):
    m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', iso_dur)
    if not m:
        return 0
    h, mn, s = m.groups()
    return int(h or 0) * 3600 + int(mn or 0) * 60 + int(s or 0)


def is_relevant(info):
    title = (info.get("title") or "").lower()
    desc = (info.get("description") or "").lower()
    text = f"{title} {desc}"
    if not any(kw.lower() in text for kw in INCLUDE_KEYWORDS):
        return False
    if any(kw.lower() in text for kw in EXCLUDE_KEYWORDS):
        return False
    dur = info.get("duration_seconds") or 0
    if dur and (dur < 25 or dur > 1200):
        return False
    if "/shorts/" in info.get("webpage_url", ""):
        return False
    return True


def classify_product(title, desc):
    text = f"{title or ''} {desc or ''}".lower()
    for kw in ["fountain", "ground fountain", "cone fountain", "ice fountain"]:
        if kw in text:
            return "fountain"
    for kw in ["cake", "barrage", "repeater", "multi-shot", "multi shot",
               "aerial cake", "500g cake", "200g cake", "compound cake"]:
        if kw in text:
            return "cake"
    return "未知"


def parse_published(published_at):
    if not published_at:
        return None, None
    try:
        dt = datetime.strptime(published_at[:19], "%Y-%m-%dT%H:%M:%S")
        return str(dt.year), f"{dt.month:02d}"
    except ValueError:
        return None, None


def create_excel_template(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "烟花视频数据"
    hf = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    b = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, b
    for col, w in {1:10, 2:45, 3:28, 4:12, 5:10, 6:10, 7:12, 8:10, 9:12,
                   10:10, 11:18, 12:24, 13:14, 14:12, 15:10, 16:12, 17:12,
                   18:12, 19:18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    wb.save(filepath)
    print(f"  Excel template created: {filepath}")


def append_data(filepath, rows):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["烟花视频数据"]
    sr = ws.max_row + 1
    df = Font(name="Arial", size=10)
    da = Alignment(vertical="center")
    uf = Font(name="Arial", size=10, color="0563C1", underline="single")
    b = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
    mf = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    uaf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for ri, rd in enumerate(rows, sr):
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = uf if ci == 2 else df
            c.alignment, c.border = da, b
            if ci in MANUAL_COLS:
                c.fill = mf
            elif ci in UNAVAILABLE_COLS:
                c.fill = uaf
    wb.save(filepath)
    print(f"  {len(rows)} data rows appended (starting row {sr})")


def collect_data(output_file):
    if not API_KEY:
        print("ERROR: YOUTUBE_API_KEY not set!")
        sys.exit(1)
    print("=" * 60)
    print("  YouTube Firework Data Collector | API v3")
    print("  cake & fountain | 2023-2026")
    print("=" * 60)
    create_excel_template(output_file)
    print("\n[Phase 1] Searching via YouTube Data API v3...")
    all_videos = {}
    total_quota = 0
    for qi, query in enumerate(SEARCH_QUERIES):
        if qi > 0:
            time.sleep(1.5)  # rate-limit between queries
        print(f"  Query: '{query[:60]}...'")
        results = search_youtube_api(query, max_results=MAX_RESULTS_PER_QUERY)
        total_quota += 100
        print(f"    -> {len(results)} results")
        for info in results:
            vid = info.get("id")
            if vid and vid not in all_videos:
                all_videos[vid] = info
    print(f"\n  Unique videos: {len(all_videos)}")
    print(f"  Quota used: {total_quota}/10000")
    print("\n[Phase 2] Filtering...")
    filtered = {vid: info for vid, info in all_videos.items() if is_relevant(info)}
    print(f"  After filtering: {len(filtered)} videos")
    print("\n[Phase 3] Fetching statistics (batch)...")
    rows = []
    sample_id = 0
    video_ids = list(filtered.keys())
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i:i+50]
        details = get_batch_details(batch)
        total_quota += 1
        for vid in batch:
            info = filtered[vid]
            detail = details.get(vid, {})
            if detail:
                info.update(detail)
            iso = info.get("duration_iso", "")
            info["duration_seconds"] = parse_iso_duration(iso) if iso else 0
            if not is_relevant(info):
                continue
            year, month = parse_published(info.get("published_at", ""))
            if year is None:
                continue
            if int(year) < 2023 or int(year) > 2026:
                continue
            title = info.get("title") or "N/A"
            desc = info.get("description") or ""
            ch = info.get("uploader") or "N/A"
            vc = info.get("view_count") or 0
            lc = info.get("like_count") or 0
            cc = info.get("comment_count") or 0
            pt = classify_product(title, desc)
            eng = round((lc + cc) / vc, 8) if vc > 0 else 0

            # Auto-fill powder weight from title/description
            powder_text = f"{title} {desc}".lower()
            powder_g = ""
            for keyword, grams in POWDER_AUTO_FILL.items():
                if keyword.lower() in powder_text:
                    powder_g = grams
                    break

            sample_id += 1
            rows.append([sample_id, f"https://www.youtube.com/watch?v={vid}",
                         ch, "", year, month, pt,
                         "", "", powder_g, "", "",
                         vc, "", "", lc, cc, "", eng])
    print(f"  Quota used: {total_quota}/10000")
    print(f"\n[Phase 4] Writing Excel: {len(rows)} samples")
    if rows:
        append_data(output_file, rows)
    else:
        print("  No data rows.")
    cakes = sum(1 for r in rows if r[6] == "cake")
    fountains = sum(1 for r in rows if r[6] == "fountain")
    print(f"\n  DONE - {len(rows)} samples (Cake: {cakes} Fountain: {fountains})")

if __name__ == "__main__":
    collect_data(str(Path(__file__).parent / OUTPUT_FILE))
